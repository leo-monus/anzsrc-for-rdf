#!/usr/bin/env python3
"""Convert an ANZSRC 2020 spreadsheet (FoR or SEO) into a SKOS RDF vocabulary.

The ABS workbooks (e.g. ``anzsrc-for_2020.xlsx`` for Fields of Research, or the
Socio-Economic Objectives workbook) each contain six worksheets. This converter
uses:

* ``Table 3`` - the master list of every concept (Divisions, Groups and
  Fields/Objectives) with its preferred label. The authoritative concept list.
* ``Table 4`` - definitions and exclusions for Divisions and Groups only.
* ``Explanatory Notes`` - free text used for the concept-scheme ``skos:note``.

``Table 1`` (Divisions) and ``Table 2`` (Divisions + Groups) are strict
subsets of ``Table 3`` and are only used for a sanity cross-check.

Hierarchy is derived from the numeric notation:

* 2-digit  -> Division -> a top concept of the scheme
* 4-digit  -> Group    -> ``skos:broader`` = its first 2 digits
* 6-digit  -> Field/Objective -> ``skos:broader`` = its first 4 digits

SKOS mapping per concept
------------------------
* ``rdf:type``          ``skos:Concept``
* ``skos:prefLabel``    preferred label (language tagged, default ``en``)
* ``skos:notation``     the numeric code
* ``skos:inScheme``     the concept scheme
* ``skos:topConceptOf`` / scheme ``skos:hasTopConcept`` (Divisions only)
* ``skos:broader`` / ``skos:narrower`` (Groups and Fields)
* ``skos:definition``   from Table 4 (Divisions and Groups)
* ``skos:scopeNote``    exclusions from Table 4 (Divisions and Groups)

Usage
-----
    python anzsrc_for_to_rdf.py \
        --xlsx /path/to/anzsrc-for_2020.xlsx \
        --output anzsrc-for_2020.rdf \
        --format pretty-xml

Run ``python anzsrc_for_to_rdf.py --help`` for all options.
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

import openpyxl
from rdflib import DCTERMS, RDF, RDFS, SKOS, Graph, Literal, Namespace, URIRef
from rdflib.namespace import XSD

# --------------------------------------------------------------------------- #
# Defaults
# --------------------------------------------------------------------------- #
DEFAULT_XLSX = "/home/u4187959/Downloads/anzsrc-for_2020.xlsx"
DEFAULT_OUTPUT = "output/anzsrc-for_2020.rdf"
DEFAULT_SCHEME_URI = "https://linked.data.gov.au/def/anzsrc-for/2020"
DEFAULT_LANG = "en"

MASTER_SHEET = "Table 3"          # all concepts + prefLabels
DEFINITION_SHEET = "Table 4"      # definitions + exclusions (Divisions/Groups)
NOTES_SHEET = "Explanatory Notes"  # scheme-level note


# --------------------------------------------------------------------------- #
# Text helpers
# --------------------------------------------------------------------------- #
def clean_label(value: object) -> str | None:
    """Collapse internal whitespace and strip a short label."""
    if value is None:
        return None
    text = str(value).replace("\xa0", " ")
    text = " ".join(text.split())
    return text or None


def clean_text(value: object) -> str | None:
    """Normalise a multi-line block (definition / exclusions).

    Non-breaking spaces are converted to regular spaces, trailing whitespace
    is stripped from each line, and surrounding blank space is removed while
    internal newlines are preserved.
    """
    if value is None:
        return None
    text = str(value).replace("\xa0", " ")
    lines = [line.rstrip() for line in text.splitlines()]
    text = "\n".join(lines).strip()
    return text or None


def as_code(value: object) -> str | None:
    """Return a clean numeric code string, or ``None`` if the cell isn't one."""
    if isinstance(value, bool):
        return None
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, str) and value.strip().isdigit():
        return value.strip()
    return None


def first_code_index(row: tuple) -> int | None:
    """Index of the first cell in ``row`` holding a numeric code."""
    for idx, cell in enumerate(row):
        if as_code(cell) is not None:
            return idx
    return None


# --------------------------------------------------------------------------- #
# Workbook parsing
# --------------------------------------------------------------------------- #
def parse_concepts(ws) -> dict[str, str]:
    """Return ``{code: prefLabel}`` from the master sheet (Table 3).

    The code sits in the first populated numeric column of a row and the label
    is in the immediately following column (the sheet is indented by level).
    """
    concepts: dict[str, str] = {}
    for row in ws.iter_rows(values_only=True):
        idx = first_code_index(row)
        if idx is None:
            continue
        code = as_code(row[idx])
        label = clean_label(row[idx + 1]) if idx + 1 < len(row) else None
        if label is None:
            print(f"  WARNING: code {code} has no label; skipping", file=sys.stderr)
            continue
        if code in concepts and concepts[code] != label:
            print(
                f"  WARNING: duplicate code {code} with differing labels "
                f"({concepts[code]!r} vs {label!r})",
                file=sys.stderr,
            )
        concepts[code] = label
    return concepts


def _find_header_columns(ws, wanted: tuple[str, ...]) -> dict[str, int]:
    """Locate named header columns (e.g. Definition/Exclusions) by scanning."""
    found: dict[str, int] = {}
    for row in ws.iter_rows(min_row=1, max_row=15, values_only=True):
        for idx, cell in enumerate(row):
            if isinstance(cell, str):
                label = cell.strip().lower()
                for name in wanted:
                    if label == name.lower():
                        found.setdefault(name, idx)
        if len(found) == len(wanted):
            break
    return found


def parse_definitions(ws) -> dict[str, dict[str, str]]:
    """Return ``{code: {"definition": ..., "scopeNote": ...}}`` from Table 4."""
    cols = _find_header_columns(ws, ("Definition", "Exclusions"))
    def_col = cols.get("Definition")
    excl_col = cols.get("Exclusions")
    if def_col is None:
        raise ValueError(f"Could not find 'Definition' column in {ws.title!r}")

    out: dict[str, dict[str, str]] = {}
    for row in ws.iter_rows(values_only=True):
        idx = first_code_index(row)
        if idx is None:
            continue
        code = as_code(row[idx])
        definition = clean_text(row[def_col]) if def_col < len(row) else None
        exclusions = (
            clean_text(row[excl_col])
            if excl_col is not None and excl_col < len(row)
            else None
        )
        entry: dict[str, str] = {}
        if definition:
            entry["definition"] = definition
        if exclusions:
            entry["scopeNote"] = exclusions
        if entry:
            out[code] = entry
    return out


def parse_scheme_note(ws) -> str | None:
    """Build the scheme-level note from the Explanatory Notes worksheet."""
    if ws is None:
        return None
    blocks: list[str] = []
    for row in ws.iter_rows(values_only=True):
        cells = [clean_text(c) for c in row]
        cells = [c for c in cells if c]
        if cells:
            blocks.append("\n".join(cells))
    return "\n\n".join(blocks) or None


def find_title(wb) -> str:
    """Compose the scheme title from the workbook header rows.

    The header block looks like::

        1297.0 ANZSRC - Australian and New Zealand Standard ..., 2020
        Fields of Research (FoR)      # or: Socio-Economic Objectives (SEO)
        Released at 11.30am ...

    so the title becomes ``"<base>: <subtitle>"`` regardless of vocabulary.
    """
    for name in (NOTES_SHEET, MASTER_SHEET, "Contents"):
        if name not in wb.sheetnames:
            continue
        rows = list(wb[name].iter_rows(min_row=1, max_row=6, values_only=True))
        parts = [clean_label(r[0]) for r in rows if r and clean_label(r[0])]
        base = next((p for p in parts if p.startswith("1297.0")), None)
        if not base:
            continue
        # Subtitle is the next header line after the base that is not the
        # release-date line ("Fields of Research (FoR)" / "Socio-Economic
        # Objectives (SEO)").
        sub = None
        for p in parts[parts.index(base) + 1:]:
            if not p.lower().startswith("released"):
                sub = p
                break
        return f"{base}: {sub}" if sub else base
    return "ANZSRC 2020"


# --------------------------------------------------------------------------- #
# Graph construction
# --------------------------------------------------------------------------- #
def build_graph(
    concepts: dict[str, str],
    definitions: dict[str, dict[str, str]],
    scheme_note: str | None,
    title: str,
    scheme_uri: str,
    lang: str,
    issued: str | None,
) -> Graph:
    g = Graph()
    g.bind("skos", SKOS)
    g.bind("dcterms", DCTERMS)
    g.bind("rdfs", RDFS)

    scheme = URIRef(scheme_uri)
    base = scheme_uri.rstrip("/") + "/"
    ns = Namespace(base)

    # Concept scheme + metadata.
    g.add((scheme, RDF.type, SKOS.ConceptScheme))
    g.add((scheme, DCTERMS.title, Literal(title, lang=lang)))
    g.add((scheme, RDFS.label, Literal(title, lang=lang)))
    g.add((scheme, SKOS.prefLabel, Literal(title, lang=lang)))
    if issued:
        g.add((scheme, DCTERMS.issued, Literal(issued, datatype=XSD.date)))
    if scheme_note:
        g.add((scheme, SKOS.note, Literal(scheme_note, lang=lang)))

    def uri(code: str) -> URIRef:
        return ns[code]

    for code, label in concepts.items():
        concept = uri(code)
        g.add((concept, RDF.type, SKOS.Concept))
        g.add((concept, SKOS.prefLabel, Literal(label, lang=lang)))
        g.add((concept, SKOS.notation, Literal(code)))
        g.add((concept, SKOS.inScheme, scheme))

        if len(code) == 2:  # Division -> top concept
            g.add((concept, SKOS.topConceptOf, scheme))
            g.add((scheme, SKOS.hasTopConcept, concept))
        else:  # Group (4) or Field (6) -> broader is the prefix
            parent = code[:-2]
            if parent in concepts:
                parent_uri = uri(parent)
                g.add((concept, SKOS.broader, parent_uri))
                g.add((parent_uri, SKOS.narrower, concept))
            else:
                print(
                    f"  WARNING: parent {parent} of {code} not found; "
                    "no broader/narrower link created",
                    file=sys.stderr,
                )

        extra = definitions.get(code)
        if extra:
            if "definition" in extra:
                g.add((concept, SKOS.definition, Literal(extra["definition"], lang=lang)))
            if "scopeNote" in extra:
                g.add((concept, SKOS.scopeNote, Literal(extra["scopeNote"], lang=lang)))

    return g


# --------------------------------------------------------------------------- #
# Validation
# --------------------------------------------------------------------------- #
def validate(concepts: dict[str, str], definitions: dict[str, dict[str, str]]) -> None:
    by_len: dict[int, int] = {}
    for code in concepts:
        by_len[len(code)] = by_len.get(len(code), 0) + 1
    print("Concept counts by notation length:", dict(sorted(by_len.items())), file=sys.stderr)
    print(f"Total concepts: {len(concepts)}", file=sys.stderr)

    orphans = [c for c in concepts if len(c) > 2 and c[:-2] not in concepts]
    if orphans:
        print(f"  WARNING: {len(orphans)} concepts without a parent: {orphans[:10]}", file=sys.stderr)

    defined = sum(1 for v in definitions.values() if "definition" in v)
    scoped = sum(1 for v in definitions.values() if "scopeNote" in v)
    print(f"Definitions: {defined}; scope notes (exclusions): {scoped}", file=sys.stderr)


# --------------------------------------------------------------------------- #
# Main
# --------------------------------------------------------------------------- #
def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--xlsx", default=DEFAULT_XLSX, help="Input .xlsx path")
    parser.add_argument("--output", default=DEFAULT_OUTPUT, help="Output RDF path")
    parser.add_argument("--scheme-uri", default=DEFAULT_SCHEME_URI, help="Concept scheme URI")
    parser.add_argument("--lang", default=DEFAULT_LANG, help="Language tag for literals")
    parser.add_argument(
        "--format",
        default="pretty-xml",
        help="rdflib serialization format (e.g. pretty-xml, xml, turtle, json-ld)",
    )
    parser.add_argument("--issued", default="2020-06-30", help="dcterms:issued date (or empty to omit)")
    args = parser.parse_args(argv)

    xlsx_path = Path(args.xlsx)
    if not xlsx_path.exists():
        parser.error(f"Input workbook not found: {xlsx_path}")

    print(f"Loading workbook: {xlsx_path}", file=sys.stderr)
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    for required in (MASTER_SHEET, DEFINITION_SHEET):
        if required not in wb.sheetnames:
            parser.error(f"Required worksheet {required!r} not found in workbook")

    concepts = parse_concepts(wb[MASTER_SHEET])
    definitions = parse_definitions(wb[DEFINITION_SHEET])
    scheme_note = parse_scheme_note(wb[NOTES_SHEET] if NOTES_SHEET in wb.sheetnames else None)
    title = find_title(wb)

    validate(concepts, definitions)

    graph = build_graph(
        concepts=concepts,
        definitions=definitions,
        scheme_note=scheme_note,
        title=title,
        scheme_uri=args.scheme_uri,
        lang=args.lang,
        issued=args.issued or None,
    )

    out_path = Path(args.output)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    graph.serialize(destination=str(out_path), format=args.format)
    print(f"Wrote {len(graph)} triples to {out_path} ({args.format})", file=sys.stderr)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
